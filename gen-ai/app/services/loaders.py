import os
import tempfile
from pathlib import Path

from langchain_core.documents import Document

# Extensions supportées → catégorie de traitement
SUPPORTED_EXTENSIONS: dict[str, str] = {
    ".pdf": "pdf",
    ".docx": "word",
    ".doc": "word",
    ".xlsx": "excel",
    ".xls": "excel",
    ".mp3": "audio",
    ".wav": "audio",
    ".m4a": "audio",
    ".ogg": "audio",
    ".flac": "audio",
    ".mp4": "video",
    ".avi": "video",
    ".mov": "video",
    ".mkv": "video",
    ".webm": "video",
}

# Modèle Whisper chargé une seule fois (lazy)
_whisper_model = None


def _get_whisper():
    global _whisper_model
    if _whisper_model is None:
        import whisper
        _whisper_model = whisper.load_model("base")
    return _whisper_model


def load_documents(file_path: str) -> list[Document]:
    ext = Path(file_path).suffix.lower()
    category = SUPPORTED_EXTENSIONS.get(ext)

    if category is None:
        raise ValueError(f"Format non supporté : {ext}")

    if category == "pdf":
        return _load_pdf(file_path)
    if category == "word":
        return _load_word(file_path)
    if category == "excel":
        return _load_excel(file_path)
    if category in ("audio", "video"):
        return _load_audio_video(file_path)

    raise ValueError(f"Catégorie inconnue : {category}")


def _load_pdf(file_path: str) -> list[Document]:
    from langchain_community.document_loaders import PyPDFLoader

    return PyPDFLoader(file_path).load()


def _load_word(file_path: str) -> list[Document]:
    from langchain_community.document_loaders import Docx2txtLoader

    return Docx2txtLoader(file_path).load()


def _load_excel(file_path: str) -> list[Document]:
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    docs: list[Document] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join("" if cell is None else str(cell) for cell in row)
            if row_text.strip():
                rows.append(row_text)
        if rows:
            docs.append(
                Document(
                    page_content="\n".join(rows),
                    metadata={"source": file_path, "sheet": sheet_name},
                )
            )

    return docs


def _load_audio_video(file_path: str) -> list[Document]:
    # Whisper gère audio et vidéo nativement via ffmpeg
    model = _get_whisper()
    result = model.transcribe(file_path)
    transcript: str = result["text"].strip()

    return [
        Document(
            page_content=transcript,
            metadata={"source": file_path},
        )
    ]
